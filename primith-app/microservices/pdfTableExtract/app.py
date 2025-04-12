import fastapi
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import Response
from gmft.auto import TableDetector, AutoTableFormatter # Assuming these exist
from gmft.pdf_bindings import PyPDFium2Document # Assuming these exist
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import pandas as pd
from urllib.parse import quote_plus
import io
import tempfile
import logging
import os
import asyncio
import gc # Garbage Collector

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - [%(threadName)s] - %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI(title="Optimized Table Extraction Service")

# Create temp directory if it doesn't exist
TEMP_DIR = os.path.join(os.getcwd(), 'temp_pdf_extract')
os.makedirs(TEMP_DIR, exist_ok=True)
logger.info(f"Temporary directory set to: {TEMP_DIR}")

# --- Enhancement 2: Optimized Excel Writing ---
def write_df_to_excel(df, sheet):
    """
    Writes a DataFrame to an Excel sheet with optimized width calculation.
    """
    df_to_write = df.copy() # Work on a copy

    if df_to_write.empty:
        logger.warning(f"DataFrame for sheet '{sheet.title}' is empty, nothing to write.")
        return

    # Optional: Check if first data row is effectively empty and skip if needed
    # (Your original logic for this can be added back here if required)
    # ...

    # Determine if columns look like default indices (0, 1, 2...)
    write_header_row = True
    try:
        current_cols = df_to_write.columns.tolist()
        expected_numeric_cols = list(range(len(current_cols)))
        # Compare as strings for robustness
        if [str(c) for c in current_cols] == [str(n) for n in expected_numeric_cols]:
            logger.info(f"Using data rows as headers for sheet '{sheet.title}'")
            write_header_row = False
    except Exception as e:
        logger.warning(f"Column header check failed for sheet '{sheet.title}': {e}")

    # --- Optimization: Track max lengths during write ---
    column_max_lengths = {} # {col_idx: max_len}

    # Write header if applicable
    if write_header_row:
        for c_idx, value in enumerate(df_to_write.columns, 1):
            cell_value = "" if pd.isna(value) else value
            cell = sheet.cell(row=1, column=c_idx, value=cell_value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Track header length
            try:
                cell_str = str(cell.value)
                lines = cell_str.splitlines()
                cell_max_line_length = max(len(line) for line in lines) if lines else 0
                column_max_lengths[c_idx] = max(column_max_lengths.get(c_idx, 0), cell_max_line_length)
            except Exception as e:
                 logger.warning(f"Error processing header length for sheet {sheet.title}, col {c_idx}: {e}")


    # Write data rows
    start_row = 2 if write_header_row else 1
    for r_idx_offset, row_data in enumerate(df_to_write.itertuples(index=False)):
        current_row_idx = start_row + r_idx_offset
        for c_idx, value in enumerate(row_data, 1):
            cell_value = "" if pd.isna(value) else value
            cell = sheet.cell(row=current_row_idx, column=c_idx, value=cell_value)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Track data length
            try:
                cell_str = str(cell.value)
                lines = cell_str.splitlines()
                cell_max_line_length = max(len(line) for line in lines) if lines else 0
                column_max_lengths[c_idx] = max(column_max_lengths.get(c_idx, 0), cell_max_line_length)
            except Exception as e:
                 logger.warning(f"Error processing cell length for sheet {sheet.title}, row {current_row_idx}, col {c_idx}: {e}")


    # Adjust column widths based on tracked max lengths
    for c_idx, max_length in column_max_lengths.items():
        try:
            column_letter = sheet.cell(row=1, column=c_idx).column_letter # Get column letter
            # Adjust width calculation slightly (padding, max width constraint)
            # Increased multiplier slightly for wrapped text, ensure minimum width
            adjusted_width = min(max( (max_length + 2) * 1.2 , 8), 70)
            sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            # Catch potential errors like invalid column index if sheet is empty/corrupt
            logger.warning(f"Could not set width for column index {c_idx} (Letter: {column_letter}) in sheet '{sheet.title}': {e}")

# --- Enhancement 1: Synchronous function for CPU-bound work ---
def _process_pdf(pdf_path):
    """
    Synchronous PDF processing function (runs in a separate thread).
    Loads models internally for simplicity in this example.
    For higher performance, use lifespan context manager to load models once.
    """
    logger.info(f"Starting PDF processing for: {pdf_path}")
    # Initialize models here for simplicity, or pass pre-initialized ones
    try:
        detector = TableDetector()
        formatter = AutoTableFormatter()
        logger.debug("Models initialized.")
    except Exception as e:
        logger.error(f"Failed to initialize models: {e}", exc_info=True)
        raise RuntimeError(f"Model initialization failed: {e}") from e

    workbook = Workbook()
    # Remove default sheet created by Workbook()
    if "Sheet" in workbook.sheetnames:
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)
        logger.debug("Removed default 'Sheet'.")

    sheets_added_count = 0
    doc = None  # Initialize doc to None

    try:
        # Manually create the document object
        doc = PyPDFium2Document(pdf_path)
        logger.info(f"Opened PDF: {pdf_path}. Processing pages...")

        # Process pages within the try block
        for page_num, page in enumerate(doc, start=1):
            logger.debug(f"Processing Page {page_num}")
            try:
                tables = detector.extract(page)
                if not tables:
                    logger.debug(f"No tables found on page {page_num}.")
                    continue
                logger.debug(f"Found {len(tables)} tables on page {page_num}.")

                for table_num, table in enumerate(tables, start=1):
                    # ... (inner try/except for table processing remains the same) ...
                    # ... (df cleaning, sheet creation, write_df_to_excel call) ...
                    logger.debug(f"Processing Page {page_num}, Table {table_num}")
                    try:
                        formatted_table = formatter.extract(table)
                        # Check if formatted_table itself is None or its df is None
                        if formatted_table is None or formatted_table.df() is None:
                            logger.warning(f"Formatter returned no data structure for table {table_num} on page {page_num}.")
                            continue

                        df = formatted_table.df() # Get DataFrame
                        if df is None or df.empty: # Check again if df is None or empty
                            logger.info(f"Empty DataFrame for table {table_num} on page {page_num}.")
                            continue

                        # Clean data more carefully
                        df = df.astype(str) # Convert all to string first
                        df.replace('None', '', inplace=True) # Replace string 'None'
                        df.replace('nan', '', inplace=True)  # Replace string 'nan'
                        # Apply newline replacement using vectorized operation if possible
                        try:
                            df = df.map(lambda x: x.replace('\\n', '\n') if isinstance(x, str) else x)
                        except Exception as replace_err:
                            logger.warning(f"Error during string replacement on P{page_num} T{table_num}: {replace_err}")


                        # Create unique sheet title
                        sheet_title_base = f"Page_{page_num}-Table_{table_num}"
                        sheet_title = sheet_title_base
                        counter = 1
                        while sheet_title in workbook.sheetnames:
                            sheet_title = f"{sheet_title_base}_{counter}"
                            counter += 1
                        
                        logger.debug(f"Creating sheet: '{sheet_title}'")
                        sheet = workbook.create_sheet(title=sheet_title)
                        write_df_to_excel(df, sheet) # Use the optimized function
                        sheets_added_count += 1
                        logger.debug(f"Finished writing to sheet: '{sheet_title}'")

                        # Hint garbage collector
                        del df
                        del formatted_table
                        gc.collect()

                    except Exception as e:
                        logger.error(f"Error processing table {table_num} on page {page_num}: {str(e)}", exc_info=True)

                # Optional: Explicitly close page object if needed by library
                # Check if the real page object has a close method
                if hasattr(page, 'close') and callable(page.close):
                     try:
                         page.close()
                         logger.debug(f"Closed page object for page {page_num}")
                     except Exception as page_close_err:
                         logger.warning(f"Could not close page object for page {page_num}: {page_close_err}")

            except Exception as page_err:
                 logger.error(f"Error processing page {page_num}: {page_err}", exc_info=True)

        logger.info(f"Finished processing all pages for {pdf_path}. Sheets added: {sheets_added_count}")
        # --- END processing loop ---

    except Exception as e:
        # Catch errors during doc opening or page iteration
        logger.error(f"Critical error processing PDF '{pdf_path}': {str(e)}", exc_info=True)
        # Let the calling function handle the HTTPException by re-raising
        # Ensure error is wrapped in a standard Exception type if needed, but RuntimeError is okay
        raise RuntimeError(f"PDF processing failed: {e}") from e
    finally:
        # --- Ensure doc.close() is always called if doc was created ---
        if doc and hasattr(doc, 'close') and callable(doc.close):
            try:
                doc.close()
                logger.info(f"Closed PDF document object for: {pdf_path}")
            except Exception as close_err:
                logger.error(f"Error closing PDF document object for {pdf_path}: {close_err}", exc_info=True)

    # Check if any sheets were added
    if sheets_added_count == 0:
        logger.warning(f"No tables were successfully extracted from {pdf_path}. Creating placeholder sheet.")
        # Ensure there's at least one sheet, even if empty
        if not workbook.sheetnames:
            workbook.create_sheet("No Tables Found")
        # If you prefer to return an error or empty response when no tables are found,
        # you could raise an exception here or return None, and handle it in the endpoint.

    # Save workbook to buffer
    excel_buffer = io.BytesIO()
    try:
        workbook.save(excel_buffer)
        logger.info(f"Excel workbook saved to buffer for {pdf_path}.")
    except Exception as save_err:
         logger.error(f"Failed to save workbook to buffer for {pdf_path}: {save_err}", exc_info=True)
         raise RuntimeError(f"Failed to save Excel file: {save_err}") from save_err
    finally:
         # Explicitly close workbook to release resources
         workbook.close()
         gc.collect()

    excel_buffer.seek(0)
    return excel_buffer.getvalue()

# --- Enhancement 1: Async function using asyncio.to_thread ---
async def process_pdf_with_timeout(pdf_path: str, timeout: int = 300):
    """
    Processes the PDF in a separate thread with a timeout.
    Manages temporary file cleanup in case of timeout or error within the thread.
    """
    logger.info(f"Scheduling PDF processing for {pdf_path} with timeout {timeout}s")
    try:
        # Runs the synchronous _process_pdf function in a thread pool managed by asyncio
        result = await asyncio.wait_for(
            asyncio.to_thread(_process_pdf, pdf_path),
            timeout=timeout
        )
        logger.info(f"Successfully processed PDF in thread: {pdf_path}")
        return result
    except asyncio.TimeoutError:
        logger.error(f"Processing timed out after {timeout} seconds for {pdf_path}")
        # No need to clean up temp file here, it's handled in the main endpoint's finally block
        raise HTTPException(status_code=504, detail=f"Processing timed out after {timeout} seconds. The PDF might be too complex or large.")
    except Exception as e:
        # Catch exceptions raised from within the thread (_process_pdf)
        logger.error(f"Error during threaded PDF processing for {pdf_path}: {str(e)}", exc_info=True)
        # No need to clean up temp file here, it's handled in the main endpoint's finally block
        # Re-raise as HTTPException for the client
        # If the error was already an HTTPException, re-raise it as is. Otherwise wrap it.
        if isinstance(e, HTTPException):
             raise e
        # Include original error type and message for better debugging
        raise HTTPException(status_code=500, detail=f"PDF processing failed: ({type(e).__name__}) {str(e)}")


@app.post("/extract-tables", response_class=Response)
async def extract_tables(file: UploadFile = File(...)):
    """
    Upload a PDF file, process it efficiently, and return an Excel file
    with extracted tables. Uses temporary file storage and threaded processing.
    """
    if not file or not file.filename or not file.filename.lower().endswith('.pdf'):
        logger.warning(f"Invalid file upload attempt: {file.filename if file else 'No file'}")
        raise HTTPException(status_code=400, detail="Invalid input: File must be a PDF.")

    temp_pdf_path = None # Initialize path variable
    try:
        # Use a temporary file to handle potentially large uploads efficiently
        # delete=False is important as we need to pass the path to another thread/process
        # and manage deletion manually after processing is complete.
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf", dir=TEMP_DIR) as temp_pdf:
            temp_pdf_path = temp_pdf.name # Store the path
            logger.info(f"Receiving file: {file.filename}. Saving to temp path: {temp_pdf_path}")
            
            # Read file content in chunks and write to temp file
            file_size = 0
            chunk_size = 1024 * 1024 # 1MB chunks
            while True:
                chunk = await file.read(chunk_size)
                if not chunk:
                    break
                file_size += len(chunk)
                temp_pdf.write(chunk)
            
            file_size_mb = file_size / (1024 * 1024)
            logger.info(f"Finished writing {file_size_mb:.2f} MB to {temp_pdf_path}")

        # Optional: Add a file size limit check here if needed
        MAX_FILE_SIZE_MB = 25 # Example limit: 100 MB
        if file_size_mb > MAX_FILE_SIZE_MB:
            logger.warning(f"File rejected: Size {file_size_mb:.2f}MB exceeds limit of {MAX_FILE_SIZE_MB}MB.")
            raise HTTPException(
                status_code=413, # Payload Too Large
                detail=f"File too large: {file_size_mb:.2f}MB exceeds limit of {MAX_FILE_SIZE_MB}MB"
            )

        # Process the temporary file using the threaded function with timeout
        processing_timeout = 600 # Example: 10 minutes timeout
        excel_content = await process_pdf_with_timeout(temp_pdf_path, timeout=processing_timeout)

        # Check if content is empty (can happen if _process_pdf decides not to create "No Tables Found")
        # Adjust this logic based on whether _process_pdf always returns bytes
        if not excel_content:
             logger.warning(f"Processing resulted in empty content for {temp_pdf_path}. Returning 204 No Content.")
             # Option: Return 204 No Content if no tables found and no placeholder sheet created
             return Response(status_code=204)

        # Prepare filename for download
        base_filename = file.filename.rsplit('.', 1)[0]
        output_filename = f"extracted_{base_filename}.xlsx"
        logger.info(f"Successfully processed {temp_pdf_path}. Sending Excel file: {output_filename}")

        # Return the generated Excel file content
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                # Ensure filename is properly encoded for headers
                "Content-Disposition": f"attachment; filename*=UTF-8''{quote_plus(output_filename)}"
            }
        )

    except HTTPException as http_exc:
        # Log HTTP exceptions specifically and re-raise them
        logger.error(f"HTTP Error during extraction for {file.filename}: Status {http_exc.status_code}, Detail: {http_exc.detail}")
        raise http_exc # Re-raise the exception to be handled by FastAPI

    except Exception as e:
        # Catch any other unexpected errors during upload or processing setup
        logger.error(f"Unhandled error in extract_tables endpoint for {file.filename}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"An unexpected server error occurred: {str(e)}")

    finally:
        # --- Crucial: Ensure temporary file is always deleted ---
        if temp_pdf_path and os.path.exists(temp_pdf_path):
            try:
                os.unlink(temp_pdf_path)
                logger.info(f"Temporary file removed: {temp_pdf_path}")
            except Exception as e:
                # Log if deletion fails, but don't crash the request handling
                logger.error(f"CRITICAL: Failed to remove temporary file {temp_pdf_path}: {e}", exc_info=True)
        # Hint garbage collection after request potentially involving large data
        gc.collect()


@app.get("/health")
async def health_check():
    """Basic health check endpoint."""
    logger.debug("Health check endpoint called.")
    return {"status": "healthy"}

@app.get("/")
async def root():
    """Root endpoint providing basic info."""
    return {"message": "Optimized Table Extraction Service. POST PDF to /extract-tables"}

# Example for running with uvicorn (if running this script directly)
if __name__ == "__main__":
    import uvicorn
    logger.info("Starting Uvicorn server...")
    # Use reload=True only for development
    # Set workers based on CPU cores for production, e.g., workers=4
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True) # Use the filename saved as main.py